import pandas as pd
import openpyxl
import zipfile
import io
import streamlit as st
from typing import Tuple, Optional

def process_marks(marks: str) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """
    Process marks string and return internal, external, and total marks.
    Handles both formats: '040+043' and '040'
    
    Args:
        marks: String containing marks in either 'internal+external' or single total format
        
    Returns:
        Tuple of (internal, external, total) marks, or (None, None, None) if invalid
    """
    if not marks:
        return None, None, None
        
    if isinstance(marks, (int, float)):
        total = int(marks)
        return None, None, total
    
    marks_str = str(marks).strip()
    
    try:
        if '+' in marks_str:
            internal, external = marks_str.split('+')
            internal_val = int(internal.lstrip('0') or '0')
            external_val = int(external.lstrip('0') or '0')
            return internal_val, external_val, internal_val + external_val
        else:
            total = int(marks_str.lstrip('0') or '0')
            return None, None, total
            
    except (ValueError, TypeError):
        return None, None, None

def process_excel_file(df: pd.DataFrame) -> openpyxl.Workbook:
    """
    Process the uploaded file (Excel or CSV) and add internal/external/total columns.
    
    Args:
        df: Input DataFrame with marksheet data
        
    Returns:
        Processed openpyxl Workbook
    """
    # Calculate the number of subjects based on the remaining columns after the first 3
    total_columns = len(df.columns)
    num_subjects = (total_columns - 3) // 4  # Assuming 4 columns per subject (Code, Name, Marks, Result)
    
    # Generate headers dynamically
    headers = ['Register No', 'Name', 'College ID']
    for i in range(1, num_subjects + 1):
        headers.extend([f'Subject Code {i}', f'Subject Name {i}', 
                        f'Marks {i}', f'Result {i}'])
    
    # If there are extra columns, add them as unnamed columns
    extra_cols = total_columns - len(headers)
    if extra_cols > 0:
        headers.extend([f'Unnamed: {i}' for i in range(extra_cols)])
    
    # Assign headers to DataFrame
    df.columns = headers
    
    # Convert to openpyxl workbook
    temp_path = 'temp_workbook.xlsx'
    df.to_excel(temp_path, index=False)
    workbook = openpyxl.load_workbook(temp_path)
    sheet = workbook.active
    
    # Process marks and insert columns
    start_col = 4  # Starting after 'Register No', 'Name', 'College ID'
    for subject_num in range(num_subjects):
        marks_col = start_col + (subject_num * 7) + 2  # Position of Marks column
        insert_col = marks_col + 1
        
        # Insert columns for Internal, External, Total
        sheet.insert_cols(insert_col, 3)
        for idx, col_name in enumerate(['Internal', 'External', 'Total']):
            sheet.cell(row=1, column=insert_col + idx, 
                      value=f'{col_name} {subject_num + 1}')
        
        # Process marks for each row
        for row in range(2, sheet.max_row + 1):
            marks = sheet.cell(row=row, column=marks_col).value
            internal, external, total = process_marks(marks)
            
            values = [
                internal if internal is not None else '',
                external if external is not None else '',
                total if total is not None else ''
            ]
            for col_offset, value in enumerate(values):
                sheet.cell(row=row, column=insert_col + col_offset, value=value)
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = max_length + 2
    
    return workbook

def create_department_batches(workbook: openpyxl.Workbook) -> io.BytesIO:
    """
    Create batch-wise Excel files for each department.
    
    Args:
        workbook: Processed workbook containing all marks
        
    Returns:
        BytesIO object containing zipped department/batch files
    """
    department_codes = {
        '28M': 'Data Science', '25F': 'BBA', '25N': 'BBAIB',
        '2AA': 'BCom', '2AK': 'BComPA', '26U': 'Psychology',
        '22S': 'Viscom', '21C': 'Economics', '21G': 'Tamil',
        '31B': 'MSW', '21B': 'Political Science', 
        '31M': 'M. Political Science'
    }
    
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    batch_workbooks = {}
    
    # Categorize rows by department and batch
    for row in range(2, sheet.max_row + 1):
        register_no = sheet.cell(row=row, column=1).value
        if not isinstance(register_no, str):
            continue  # Skip rows with invalid register numbers
        dept_code = register_no[2:5]
        batch_year = register_no[:2]
        
        if dept_code in department_codes:
            dept_name = department_codes[dept_code]
            row_values = [cell.value for cell in sheet[row]]
            
            if dept_name not in batch_workbooks:
                batch_workbooks[dept_name] = {}
                
            if batch_year not in batch_workbooks[dept_name]:
                new_wb = openpyxl.Workbook()
                new_wb.active.append(headers)
                batch_workbooks[dept_name][batch_year] = new_wb
                
            batch_workbooks[dept_name][batch_year].active.append(row_values)
    
    # Create ZIP file with all workbooks
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for dept, batches in batch_workbooks.items():
            for batch_year, wb in batches.items():
                batch_file = f'{dept.replace(" ", "_")}_Batch_{batch_year}.xlsx'
                wb.save(batch_file)
                zip_file.write(batch_file, batch_file)
                import os
                os.remove(batch_file)
    
    zip_buffer.seek(0)
    return zip_buffer

def main():
    st.title("Marksheet Processing and Department-wise Excel Export")
    
    # Allow both Excel and CSV files
    uploaded_file = st.file_uploader("Upload a Marksheet file", type=["xlsx", "csv"])
    
    if uploaded_file is not None:
        try:
            # Determine file type and read accordingly
            file_extension = uploaded_file.name.split('.')[-1].lower()
            
            if file_extension == 'xlsx':
                df = pd.read_excel(uploaded_file, header=None)
            elif file_extension == 'csv':
                df = pd.read_csv(uploaded_file, header=None)
            else:
                st.error("Unsupported file format. Please upload an .xlsx or .csv file.")
                return
            
            # Process the file
            processed_workbook = process_excel_file(df)
            
            # Create department/batch-wise files
            zip_buffer = create_department_batches(processed_workbook)
            
            # Provide download button
            st.download_button(
                label="Download Department and Batch-wise Excel Files (ZIP)",
                data=zip_buffer,
                file_name="department_batch_excel_files.zip",
                mime="application/zip"
            )
            
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()
